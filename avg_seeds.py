#! /usr/bin/env python
""" Make standard average output sheets from Trarr simulations.

By Neale Irons version 06/04/2019 (CC BY-SA 4.0)
todo: resize and centre columns, font to Arial 10
"""

from __future__ import print_function
from itertools import tee, zip_longest
from openpyxl.styles import Font
from openpyxl.styles import Font, colors
from openpyxl.utils import column_index_from_string
from sys import argv
import argparse
import fnmatch
import glob
import openpyxl
import os
import sys


def usage():
    """ Display usage information """
    parser = argparse.ArgumentParser(description=
            "Make standard excel average sheets from Trarr output files.",
            epilog=' eg: avg_seed HW2S3*')
    parser.add_argument('filename', nargs='?',
                        help='full or part file specification (default: all)')
    args = parser.parse_args()


def finalise_sheet(wb, output, nfile, length, sum_unimpeded_speed, basefile):
    """ Calculate averages and write file """
    output, unimpeded_speed = calculate_averages(output, nfile,
                                                 sum_unimpeded_speed)
    write_average_sheet(wb, output, nfile, length, unimpeded_speed)
    wb.save(str(basefile) + '_AVERAGES.xls')


def calculate_averages(output, nfile, sum_unimpeded_speed):
    """ Average output and unimpeded speed """
    # Divide array_sum by number seed files
    for i in range(len(BLK)):
        for j in range(BLOCK_DATA_ROWS - 2):
            for k in range(len(INDICIES)):
                if output[0][i][j][k] > 0:
                    output[0][i][j][k] = round(output[0][i][j][k] /
                                            output[1][i][j][k], 1)
                # print(output[0][i][j][k], end = " ")    # Debug
            # print()                                     # Debug
    unimpeded_speed = round(sum_unimpeded_speed / nfile, 1)
    return(output, unimpeded_speed)


def write_average_sheet(wb, output, nfile, length, unimpeded_speed):
    """ Write summary sheet in compatible format """
    ws = wb.create_sheet("Average")
    for sheet in range(len(wb.sheetnames)):
        if wb.sheetnames[sheet] == "Average":
            break
    wb.active = sheet
    ws['A2'] = nfile
    ws['A2'].font = Font(bold=True)
    ws['B2'] = "Seed files averaged"
    ws['B2'].font = Font(bold=True)
    ws['A5'] = ROW_HDR[0][0]
    ws['B5'] = length

    # For each block
    for bidx in range(len(BLK)):
        # Upper block
        row = 9
        xcol = column_index_from_string(BLK[bidx][4])
        ws.cell(row=row, column=xcol).value = BLK[bidx][0]
        ws.cell(row=row, column=xcol).font = Font(bold=True)
        row += 1
        # Column headings
        for col in range(0, 6):
            ws.cell(row=row, column=xcol + col).value = COL_HDR[col][0]
        # Row details
        xrow = 10
        for row in range(1, 8):
            idx = 0
            ws.cell(row=xrow + row, column=xcol + idx).value = ROW_HDR[row][0]
            # Selected columns from array
            for col in (2, 4, 12, 13, 6):
                idx += 1
                # print(bidx, xrow, row, xcol, col)  # Debug
                ws.cell(row=xrow + row,
                        column=xcol + idx).value = output[0][bidx][row][col-1]
        # Lower block
        xrow = 19
        ws.cell(row=xrow, column=xcol).value = COL_HDR[0][0]
        ws.cell(row=xrow, column=xcol + 1).value = COL_HDR[6][0]
        for row in range(1, 8):
            ws.cell(row=xrow + row, column=xcol).value = ROW_HDR[row][0]
            ws.cell(row=xrow + row,
                    column=xcol + 1).value = round(output[0][bidx][row][15-1])

    # Footer details
    xrow = 29
    for row in range(0, 8):
        ws.cell(row=xrow + row, column=2).value = ROW_HDR[row + 8][0]
        if row == 0:
            ws.cell(row=xrow + row, column=2).font = Font(bold=True)
        elif row == 1:
            ws.cell(row=xrow + row, column=3).value = "=C17"
            ws.cell(row=xrow + row, column=4).value = "=IF(C30>90,1,IF(C30>80,2,IF(C30>70,3,IF(C30>60,4,5))))"
            ws.cell(row=xrow + row, column=4).font = Font(color="FF969696")
        elif row == 2:
            ws.cell(row=xrow + row, column=3).value = "=F17"
            ws.cell(row=xrow + row, column=4).value = "=IF(C31>80,5,IF(C31>65,4,IF(C31>50,3,IF(C31>35,2,1))))"
            ws.cell(row=xrow + row, column=4).font = Font(color="FF969696")
        elif row == 3:
            ws.cell(row=xrow + row, column=3).value = unimpeded_speed
            ws.cell(row=xrow + row, column=4).value = "=MAX(D30:D31)"
            ws.cell(row=xrow + row, column=4).font = Font(color="FF969696")
            xrow += 2  # Blank rows
        elif row == 4:
            ws.cell(row=xrow + row, column=2).font = Font(bold=True)
            ws.cell(row=xrow + row, column=3).value = COL_HDR[7][0]
            ws.cell(row=xrow + row, column=3).font = Font(bold=True)
        elif row == 5:
            ws.cell(row=xrow + row, column=3).value = '=IF(D32=1,"A",IF(D32=2,"B",IF(D32=3,"C",IF(D32=4,"D","E"))))'
        elif row == 6:
            ws.cell(row=xrow + row, column=3).value = '=IF(C31>85,"E",IF(C31>70,"D",IF(C31>55,"C",IF(C31>40,"B","A"))))'
        elif row == 7:
            ws.cell(row=xrow + row, column=3).value = '=IF(C32>91.7,"A",IF(C32>83.3,"B",IF(C32>75,"C",IF(C32>66.7,"D","E"))))'


def main():

    global BLOCK_DATA_ROWS
    global BLOCK_LAYOUT_ROWS
    global INDICIES
    global BLK
    global COL_HDR
    global ROW_HDR

    BLOCK_DATA_ROWS = 10
    BLOCK_LAYOUT_ROWS = 13
    INDICIES = [0, 9, 17, 23, 33, 38, 46, 54, 65, 72, 82, 91, 100, 110, 119]
    #      ['col_hdr',         'search',   'extra rows', 'out_col', 'av_col]
    BLK = [['Direction 1',     b' DIR1',     2,           'R',       'H'],
           ['Direction 2',     b' DIR2',     2,           'AI',      'O'],
           ['Both Directions',
                    b'          * INTERVAL', 3,           'A',       'A']]
    #      ['unimpeded_speed', ' ** FREE',  13, 21:27,   'F']
    #      ['Seed',            '_seed',     0,           'A']

    COL_HDR = [['Vehicle Category'],
               ['Travel Time (sec)'],
               ['Speed (km/hr)'],
               ['Petrol Cons ML'],
               ['Diesel Cons ML'],
               ['% Time Spent Foll'],
               ['Flow (v/hr)'],
               ['LOS']]

    ROW_HDR = [['Length (m)'],
               ['Cars'],
               ['Cars Towing'],
               ['Rigids'],
               ['Single Artics'],
               ['Double Artics'],
               ['Road Trains'],
               ['All'],
               ['All Vehicle Averages'],
               ['Speed'],
               ['PTSF'],
               ['Unimpeded Speed'],
               ['HCM Highway Class'],
               ['If Class I Hwy'],
               ['If Class II Hwy'],
               ['If Class III Hwy']]

    dirlist = os.listdir('.')
    pattern = '*.OUT'
    if len(argv) > 1:
        # File integrity check
        if next(glob.iglob(argv[1]), None):
            pattern = argv[1] + '.OUT'
        else:
            if not '-h' in argv[1]:
                print("File path %s does not exist." % argv[1])
            usage()
            sys.exit()
    basefile = ""

    for file in dirlist:
        if fnmatch.fnmatch(file, pattern):
            if not basefile or basefile != str(file.split('_seed')[0]):
                if basefile and basefile != str(file.split('_seed')[0]):
                    # Calculate averages and write file
                    finalise_sheet(wb, output, nfile, length,
                                   sum_unimpeded_speed, basefile)
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = 'Model_Output'
                output = [[[[0.0 for i in range(len(INDICIES))]     # Data
                             for j in range(BLOCK_DATA_ROWS - 2)]   # Dir
                             for k in range(len(BLK))]              # Vehicle
                             for n in range(2)]                     # Totals
                nfile = 0
                seed = ""
                length = 0
                sum_unimpeded_speed = 0.0
                basefile = str(file.split('_seed')[0])
                print(basefile,)
            seed = int(str(file.split('_seed')[1].split('.')[0]))
            if seed <= 0 or seed > 999999:
                sys.exit("Invalid filename %s - seed must be between 0 and"
                         " 999999. Exiting..." % file)

            for bidx in range(len(BLK)):
                # Get extra column offset for block
                xcol = column_index_from_string(BLK[bidx][3])
                ws.cell(row=3, column=xcol).value = BLK[bidx][0]
                ws.cell(row=3, column=xcol).font = Font(bold=True)
            df = file
            with open(df, 'rb') as f:
                nfile += 1
                # Get extra row offset for block
                xrow = BLOCK_LAYOUT_ROWS * nfile
                # For each block ***
                for bidx in range(len(BLK)):
                    # print("Processing: ", nfile, BLK[bidx][0], xrow)  # Debug
                    xcol = column_index_from_string(BLK[bidx][3])
                    if not bidx:
                        # Skip file header
                        for index, line in enumerate(f):
                            if index == 2:
                                break
                        # Check seed
                        if seed != int(line.split(b'.')[0].split(b'_seed')[1]):
                            print(seed,  int(line.split(b'.')[0].split(b'_seed')[1]))
                            seed = line.split(b'.')[0].split(b'_seed')[1]
                            sys.exit("Filename %s does not contain seed."
                                     " Exiting..." % file)
                    # Find block
                    for index, line in enumerate(f):
                        if line.startswith(BLK[bidx][1]):
                            # print(length)
                            if length == 0:
                                length = (int(line.split(b'(')[1].split(b'.')[0]))
                            break
                    # Skip block header
                    for line in range(BLK[bidx][2] - 1):
                        next(f)
                    # write seed
                    ws.cell(row=xrow - 1, column=xcol).value = seed
                    # Load each model output row line from block
                    for index, line in enumerate(f):
                        if index == BLOCK_DATA_ROWS:
                            break
                        start, end = tee(INDICIES)
                        next(end)
                        line = line.rstrip(b'\r\n')
                        fields = [line[i:j].strip(b' ') for i, j in zip_longest(start, end)]

                        for col, field in enumerate(fields, start=0):
                            ws.cell(row=index + xrow, column=col + xcol).value = field
                        if index > 2:  # Number of header lines - 1
                            for k in range(1, len(INDICIES)):  # Data for each field
                                if fields[k] != b'':
                                    output[0][bidx][index-2][k] += float(fields[k])
                                    if float(fields[k]) > 0:
                                        output[1][bidx][index-2][k] += 1

                # Find unimpeded speed all
                for index, line in enumerate(f):
                    if line.startswith(b' ** FREE'):
                        break
                # Skip block header
                for index, line in enumerate(f):
                    if index == 12:
                        break
                # Write
                ws.cell(row=xrow - 1, column=3).value = "Unimpeded Speed All Vehicles"
                ws.cell(row=xrow - 1, column=6).value = float(line[20:26])
                sum_unimpeded_speed += float(line[20:26])

    # Last file
    if basefile:
        finalise_sheet(wb, output, nfile, length, sum_unimpeded_speed,
                       basefile)
        print(" done.")
    else:
        print(" No files matched - none processed.")


if __name__ == '__main__':
    main()
