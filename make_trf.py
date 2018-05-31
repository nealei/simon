#! /usr/bin/env python
""" Make standard average output sheets from Trarr simulations.

Program requires template.TFR in same directory and
 trf files created in same directory as trf_values workbook.

By Neale Irons version 30/05/2018 (CC BY-SA 4.0)
"""

from __future__ import print_function
import argparse
import fnmatch
import openpyxl
import os
import shutil
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from shutil import move
from sys import argv


def usage():
    """ Display usage information """
    parser = argparse.ArgumentParser(description=
            "Make Trarr TRF files from traffic survey reports.",
            epilog=' eg: make_trf.py "Trarr trf Values Demo.xlsm"'
                   ' Filenames with spaces must be enclosed in quotes.')
    parser.add_argument('filename', help='Excel workbook to process')
    args = parser.parse_args()


def write_traff_files(ws, src_name, col):
    """ Write traff files for each seed. """
    TRF_NAME_ROW = 3
    dst_basename = ws.cell(column=col, row=TRF_NAME_ROW).value  # Base f'name
    print(dst_basename)
    for seed in enabled_seeds(ws, col):
        # print(seed)
        dst_name = dst_basename + "_seed" + str(seed) + ".TRF"
        with open(src_name) as src, open(dst_name, 'w') as dst:  # Temp file
            templines = src.read().splitlines()
            templines[0] = dst_name
            padded_seed = "      " + str(seed)  # Works
            templines[18] = padded_seed[-6:] + templines[18][6:]
            # for line in templines:
            dst.writelines("%s\n" % l for l in templines)
            dst.write('\n')
            shutil.copyfileobj(src, dst)        # Copy rest of file
        # os.rename(dst.name, text_filename)      # rename new version


def get_cells(ws, row, col, no_rows):
    """ Return list of cell range values """
    for col in ws.iter_cols(min_row = row, min_col = col,
                            max_row = row + no_rows - 1, max_col = col):
        for cell in col:
            yield cell.value


def enabled_seeds(ws, col):
    """ Return list of enabled seeds from worksheet. """
    NO_SEEDS = 10
    ENABLE_START_ROW = 77
    ENABLE_END_ROW = ENABLE_START_ROW + NO_SEEDS - 1
    SEED_START_ROW = 88
    SEED_END_ROW = SEED_START_ROW + NO_SEEDS - 1

    col_letter = get_column_letter(col)
    seeds_range = [(col_letter + str(ENABLE_START_ROW) + ":" +
                    col_letter + str(ENABLE_END_ROW))]
    seeds_range.append(col_letter + str(SEED_START_ROW) + ":" +
                       col_letter + str(SEED_END_ROW))
    # print(seeds_range)      # Debug
    seeds = [cell[1][0].value for cell in zip(ws[seeds_range[0]], ws[seeds_range[1]])
                                              if cell[0][0].value]

    # Test this ***
    # use itertools.izip to zip the generators in order to process
    # the elements one by one instead of generating them all at once
    # seeds = [cell[1] for cell in itertools.izip(
    #             get_cells(ENABLE_START_ROW, col, NO_SEEDS),
    #             get_cells(SEED_START_ROW, col, NO_SEEDS)) if cell[0]]

    return seeds;


def main():

    #       [ item,  line, rw, i,  fmt   ]
    item = [['TUN',     6, 36, 0, '%8i'  ],
            ['TSE',     7, 37, 0, '%8.1f'],
            ['tsi',     8, 38, 0, '%8.1f'],
            ['OPTION', 10, 39, 0, '%8i'  ],
            ['DTS1',   12, 40, 0, '%8.1f'],
            ['DTS2',   13, 41, 0, '%8.1f'],
            ['ppfol1', 14, 42, 0, '%8.1f'],
            ['ppfol2', 15, 43, 0, '%8.1f'],
            ['NSTR',   17, 44, 0, '%8i'  ],
            ['NSEED0', 19, 45, 0, '%8.1f'],
            ['ICHECK', 20, 46, 0, '%8i'  ]]

    # DATA = [['C69:C74', 'C51:C56', 'C59:C64'],  # Source range
    DATA = [[69, 51, 59],                 # Source row start
            ['%8.1f', '%8.2f', '%8.2f']]  # Format

    TRF_CLASSES = 6
    TRF_NAME_ROW = 3                    # *** Required in other procedures
    FIRST_TRAF_COL = 3

    # File integrity check
    if len(argv) == 2 and os.path.isfile(argv[1]):
        wb_name = argv[1]
    else:
        if len(argv) > 1:
            if not '-h' in argv[1]:
                print("File path %s does not exist." % argv[1])
        usage()
        sys.exit()

    # src_name = os.path.splitext(argv[1])
    wb = load_workbook(wb_name, data_only=True)
    ws = wb['TRFdata']

    print("Processing", wb_name, "to column",
          get_column_letter(ws.max_column), "...")
    for col in range(FIRST_TRAF_COL, ws.max_column + 1):  # If unreliable use 40 ***
        for i in range(len(item)):
            item[i][TRF_NAME_ROW] = ws.cell(column=col, row=i + 36).value
        if not ws.cell(column=col, row=TRF_NAME_ROW).value:
            continue
        with open('../template.TRF') as f:
            templines = f.read().splitlines()       # Trap no file found error
        tmp_name = '../temp.trf'
        with open(tmp_name, 'w') as f:
            # i is line, j is item number, k is direction
            j = 0
            k = 1
            for i in range(len(templines)):
                if i + 1 in (6, 7, 8, 10, 12, 13, 14, 15, 17, 19, 20):
                    f.write(item[j][4] % item[j][3])
                    f.write(templines[i][8:])
                    j += 1
                elif i + 1 in (48, 50, 54):  # Vehicle proportion / volume
                    if i + 1 == 48:
                        src_row = DATA[0][k]
                    elif i + 1 == 50:
                        src_row = DATA[0][k]
                    elif i + 1 == 54:
                        src_row = DATA[0][k]
                    for element in get_cells(ws, src_row, col, TRF_CLASSES):
                        f.write(DATA[1][k] % element)
                        # print(DATA[1][k] % element)
                    f.write(templines[i][48:])
                    k = (k + 1) % 3
                else:
                    f.write(templines[i])
                f.write('\n')

        write_traff_files(ws, tmp_name, col)
    os.remove(tmp_name)
    print(" done.")


if __name__ == '__main__':
    main()
